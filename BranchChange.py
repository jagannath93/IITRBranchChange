'''
Script to automate the process of allocating new branches to sophomores at IIT Roorkee based on their preferences.
'''

# openpyxl imports
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter

# Tkinter imports
from Tkinter import *
import tkMessageBox

# Python imports
import operator
import os.path
import sys


# Global variables/data structures
statusStr = ''
btn1 = 0
btn2 = 0
pr_call_no = 0
no_of_applied_students = 0    # No of students in original BranchChange list which contains both Eligible/Non-eligible students.
no_of_eligible_students = 0  # No of students eligible for BranchChange programme
eligible_progs = []   # List of eligible programmes for Branch Change Process
OV_list = dict()
AV_list = dict()
student_by_enrno = dict()
students = []     # Eligible students for branch change
programme_ratings = dict()
ineligible_candidates = dict()  # Stores details of candidate, reson for ineligibility 
allotment_status = dict()     # Stores the status of every eligible applicant participated in allotment process
final_alloted_list_GEN = []     # Stores name, old programme, new programme alloted aganist enrollment.no
final_alloted_list_OBC = []
final_alloted_list_SC = []
final_alloted_list_ST = []


################  DEFINITIONS  ###################

def load_eligible_progs():    	    # Returns the list of eligible programmes for branch change.
  try:
    f = open('InputFiles/EligibleProgrammes.txt', 'r')      # Fetches data from 'EligibleProgrammes.txt'	
    for line in f.readlines():
      item = line.replace('\n', '')
      if not item == '':
        eligible_progs.append(item)
        programme_ratings[item] = 0
    f.close()
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured while processing eligible programmes: \n(%s)" % str(e)
    )        
    sys.exit(1)
  #global statusStr
  #print "Eligible programmes loaded"
  #Vstr.set(statusStr)

def is_prog_eligible(prog):
  if prog in eligible_progs:
    return True
  else:
    return False

def load_vacancies():
  try:
    wb = load_workbook(filename = "InputFiles/OldVacancies.xlsx")
    ws = wb.get_sheet_by_name("vacancies")
    #print len(ws.rows)
    rows = list(ws.rows)
    for row in rows:
      tmp = []
      tmp.append(int(row[1].value))
      tmp.append(int(row[2].value))
      tmp.append(int(row[3].value))
      tmp.append(int(row[4].value))
      OV_list[row[0].value] = tmp
      _total_OV = tmp[0]+tmp[1]+tmp[2]+tmp[3]
      if _total_OV == 0:
        AV_list[row[0].value] = 2
      elif _total_OV == 1:
        AV_list[row[0].value] = 1
      elif _total_OV >= 2:
        AV_list[row[0].value] = 0
        #print OV_list
        #print AV_list
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured while processing vacancies: \n(%s)" % str(e)
    )        
    sys.exit(1)
  #global statusStr
  #print "Old vacancies and additional seats loaded"
  #Vstr.set(statusStr)
   
 
def get_OV(prog):
  return int(OV_list[prog])

def get_OV_by_cat(prog, cat):
  if cat == "GE":
    return int(OV_list[prog][0])
  elif cat == "OBC":
    return int(OV_list[prog][1])
  elif cat == "SC":
    return int(OV_list[prog][2])
  elif cat == "ST":
    return int(OV_list[prog][3])
  else:
    return False

def get_AV(prog):
  return int(AV_list[prog])

def eligible_candidates_list():
  try:
    #global statusStr
    #print "Filtering eligible candidates from the entire applied candidated list..."
    #Vstr.set(statusStr)

    wb = load_workbook(filename = "InputFiles/BranchChange.xlsx")
    ws = wb.get_sheet_by_name("common")
    tmp = ws.columns
    #print len(tmp)
    rows = len(tmp[0]) - 1
    #print rows
    global no_of_applied_students
    no_of_applied_students = rows

    wb_new = Workbook()
    dest_filename = r'OutputFiles/EligibleStudents.xlsx'
    ws_new = wb_new.worksheets[0]
    ws_new.title = "common"
    for col_idx in xrange(1,11):
      col = get_column_letter(col_idx)
      row_1 = 1
      row_2 = 1
      while row_2 < rows+1:
        enrno = ws.cell('%s%s'%('A', row_2 + 1)).value
        name = ws.cell('%s%s'%('B', row_2 + 1)).value
        if ws.cell('%s%s'%('I', row_2 + 1)).value == 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == False \
               and ws.cell('%s%s'%('C', row_2 + 1)).value in eligible_progs:
          ws_new.cell('%s%s'%(col, row_1)).value = ws.cell('%s%s'%(col, row_2 + 1)).value
          row_1 = row_1 + 1
        elif ws.cell('%s%s'%('I', row_2 + 1)).value is not 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == False \
               and ws.cell('%s%s'%('C', row_2 + 1)).value in eligible_progs:
          reason = " Insufficient earned credits         "
          ineligible_candidates[ enrno ] = [name, reason]
        elif ws.cell('%s%s'%('I', row_2 + 1)).value == 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == True \
               and ws.cell('%s%s'%('C', row_2 + 1)).value in eligible_progs:
          reason = " Penality for indicipline               "
          ineligible_candidates[ enrno ] = [name, reason]
        elif ws.cell('%s%s'%('I', row_2 + 1)).value == 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == False \
               and ws.cell('%s%s'%('C', row_2 + 1)).value not in eligible_progs:
          reason = " Branch not suitable to change  "
          ineligible_candidates[ enrno ] = [name, reason]
        elif ws.cell('%s%s'%('I', row_2 + 1)).value is not 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == True \
               and ws.cell('%s%s'%('C', row_2 + 1)).value in eligible_progs:
          reason = " ID/IEC                                           "
          ineligible_candidates[ enrno ] = [name, reason]
        elif ws.cell('%s%s'%('I', row_2 + 1)).value == 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == True \
               and ws.cell('%s%s'%('C', row_2 + 1)).value not in eligible_progs:
          reason = " CBNS/ID                                        "
          ineligible_candidates[ enrno ] = [name, reason]
        elif ws.cell('%s%s'%('I', row_2 + 1)).value is not 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == False \
               and ws.cell('%s%s'%('C', row_2 + 1)).value not in eligible_progs:
          reason = " CBNS/IEC                                     "
          ineligible_candidates[ enrno ] = [name, reason]
        elif ws.cell('%s%s'%('I', row_2 + 1)).value is not 21.0 and ws.cell('%s%s'%('J', row_2 + 1)).value == True \
               and ws.cell('%s%s'%('C', row_2 + 1)).value not in eligible_progs:
          reason = " CBNS/ID/IEC                                "
          ineligible_candidates[ enrno ] = [name, reason]
        row_2 = row_2 + 1
    wb_new.save(filename = dest_filename)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured while filtering eligible candidates: \n(%s)" % str(e)
    )        
    sys.exit(1)
  #global statusStr
  #print "Eligible candidates list prepared and saved to 'EligibleStudents.xlsx'"
  #Vstr.set(statusStr)

def load_student_details():
  try:
    wb = load_workbook(filename = "OutputFiles/EligibleStudents.xlsx")
    ws = wb.get_sheet_by_name("common")			
    cols = ws.columns
    rows = list(ws.rows)
    enrnos = list(cols[0])
    global no_of_eligible_students
    no_of_eligible_students = len(rows)
    
    for row in rows:
      tmp = []
      enrno = ""
      enrno = row[0].value
      tmp.append(row[0].value)
      tmp.append(row[1].value)
      tmp.append(row[2].value)
      tmp.append(row[3].value)
      tmp.append(row[5].value)
      _var = []
      _tmp = row[6].value
      _var = _tmp.split('/')
      tmp.append(_var)
      tmp.append(row[7].value)
      tmp.append(row[8].value)
      tmp.append(row[9].value)
      student_by_enrno[enrno] = tmp
      students.append(tmp)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured while loading eligible student details: \n(%s)" % str(e)
    )        
    sys.exit(1)
  #global statusStr
  #print "Candidates details and data loaded"
  #Vstr.set(statusStr)

def get_student_details( enrno ):
  return student_by_enrno[enrno]

def programme_demand_ratings():
  for s in students:
    for prog in s[5]:
      if is_prog_eligible(prog):
        programme_ratings[prog] += 1
				
  #print programme_ratings
  #return sorted(programme_ratings, key=programme_ratings.get, reverse=True)
  #return programme_ratings

def allotment():     # Main function which carries out the allotment process in a specific category.
  try:
    #global statusStr
    #print "Starting the allotment process..."
    #Vstr.set(statusStr)
		
    for student in students:
      enrno = student[0]
      name = student[1]
      curr_programme = student[2]
      category = student[3]
      jee_rank = student[4]
      choices = student[5]
      cgpa = student[6]
      #allotment_status[enrno] = []   # Initialization

      count = 0
      vacancies = 0
      choices_no = len(choices)
      for choice in choices:
        if not is_prog_eligible(choice):
          count = count + 1
          if count == choices_no:
            helper = []
            helper.append(name)
            helper.append("Not Alloted")
            allotment_status[enrno] = helper
          continue
        else:
          if get_OV_by_cat(choice, category) > 0:
            tmp = []
            tmp.append(enrno)
            tmp.append(name)
            tmp.append(curr_programme)
            tmp.append(choice)
            if category == "GE":
              final_alloted_list_GEN.append(tmp)
              OV_list[choice][0] -= 1
            elif category == "OBC":
              final_alloted_list_OBC.append(tmp)
              OV_list[choice][1] -= 1
            elif category == "SC":
              final_alloted_list_SC.append(tmp)
              OV_list[choice][2] -= 1
            elif category == "ST":
              final_alloted_list_ST.append(tmp)
              OV_list[choice][3] -= 1
            helper = []
            helper.append(name)
            helper.append("Alloted        ")
            allotment_status[enrno] = helper
            break
          elif get_AV(choice) > 0:
            AV_list[choice] -= 1
            tmp = []
            tmp.append(enrno)
            tmp.append(name)
            tmp.append(curr_programme)
            tmp.append(choice)
            if category == "GE":
              final_alloted_list_GEN.append(tmp)
            elif category == "OBC":
              final_alloted_list_OBC.append(tmp)
            elif category == "SC":
              final_alloted_list_SC.append(tmp)
            elif category == "ST":
              final_alloted_list_ST.append(tmp)
            helper = []
            helper.append(name)
            helper.append("Alloted        ")
            allotment_status[enrno] = helper
            break
        count = count + 1
        #print str(enrno) +"     "+ str(count)+ "     "+ str(choices_no)
        if count > 0 and count == choices_no:
          helper = []
          helper.append(name)
          helper.append("Not Alloted")
          allotment_status[enrno] = helper
        elif count > 0 and count < choices_no:
          continue
        else:
          tkMessageBox.showerror(
            "Error Message",
            "Error occured while processing '"+ name + "' preference order !! "
          )        
          sys.exit(1)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured while loading eligible student details: \n(%s)" % str(e)
    )        
    sys.exit(1)

def print_result_to_excel():
  try:
    wb = Workbook()
    dest_filename = r'OutputFiles/BranchChange_result.xlsx'
    ws = wb.worksheets[0]
    ws.title = "General"
    no_of_GEN_students_alloted = len(final_alloted_list_GEN)

    ws.cell('A1').value = "Enrollment no."
    ws.cell('B1').value = "Name"
    ws.cell('C1').value = "Old Programme"
    ws.cell('D1').value = "New Programme"
    #ws.cell('E1').value = "CGPA"
    #ws.cell('F1').value = "JEE Rank"
    for row in xrange(0, no_of_GEN_students_alloted):
      for col in xrange(0, 4):
        #print str(row) +", "+ str(col)
        ws.cell(row = row+2, column = col).value = final_alloted_list_GEN[row][col]
    
    ws = wb.create_sheet()
    ws.title = "OBC"
    no_of_OBC_students_alloted = len(final_alloted_list_OBC)

    ws.cell('A1').value = "Enrollment no."
    ws.cell('B1').value = "Name"
    ws.cell('C1').value = "Old Programme"
    ws.cell('D1').value = "New Programme"
    #ws.cell('E1').value = "CGPA"
    #ws.cell('F1').value = "JEE Rank"
    for row in xrange(0, no_of_OBC_students_alloted):
      for col in xrange(0, 4):
        ws.cell(row = row+2, column = col).value = final_alloted_list_OBC[row][col]
    
    ws = wb.create_sheet()
    ws.title = "SC"
    no_of_SC_students_alloted = len(final_alloted_list_SC)

    ws.cell('A1').value = "Enrollment no."
    ws.cell('B1').value = "Name"
    ws.cell('C1').value = "Old Programme"
    ws.cell('D1').value = "New Programme"
    #ws.cell('E1').value = "CGPA"
    #ws.cell('F1').value = "JEE Rank"
    for row in xrange(0, no_of_SC_students_alloted):
      for col in xrange(0, 4):
        ws.cell(row = row+2, column = col).value = final_alloted_list_SC[row][col]

    ws = wb.create_sheet()
    ws.title = "ST"
    no_of_ST_students_alloted = len(final_alloted_list_ST)

    ws.cell('A1').value = "Enrollment no."
    ws.cell('B1').value = "Name"
    ws.cell('C1').value = "Old Programme"
    ws.cell('D1').value = "New Programme"
    #ws.cell('E1').value = "CGPA"
    #ws.cell('F1').value = "JEE Rank"
    for row in xrange(0, no_of_ST_students_alloted):
      #col = get_column_letter(col_idx+1)
      for col in xrange(0, 4):
        ws.cell(row = row+2, column = col).value = final_alloted_list_ST[row][col]

    wb.save(filename = dest_filename)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured while printing allotment result to 'BranchChange_result.xlsx' file: \n(%s)" % str(e)
    )        
    sys.exit(1)
  #global statusStr
  #print "Printing complete."
  #Vstr.set(statusStr)
              
  # Tkinter Declarations/Initializations

def callback():
  if tkMessageBox.askokcancel("Quit", "Do you really wish to quit?"):
    master.destroy()

def checkInputs():
  global L1, L2, btn1
  file1 = "InputFiles/EligibleProgrammes.txt"
  file2 = "InputFiles/OldVacancies.xlsx"
  file3 =  "InputFiles/BranchChange.xlsx"
  tester = 0

  if os.path.isfile(file1):
    tester += 1
  if os.path.isfile(file2):
    tester += 1
  if os.path.isfile(file3):
    tester += 1
  
  if tester is 3:
    L1 = Label(master, text="Detected correct input")
    L1.pack()
    btn1 = 1
    b2.config (state=ACTIVE)
  else:
    #lb.delete(0,END)
    L2 = Label(master, text="Invalid Inputs! One or More input files missing")
    L2.pack()

def index():
  try:
    global master, lb, b1, b2, b3, b4
    master = Tk()
    master.protocol("WM_DELETE_WINDOW", callback)
    frame1 = Frame(master, pady=30)
    frame1.pack()
    Label(frame1, text="Branch Change Process Automator", font=("Helvetica", 12)).pack()
    frame2 = Frame(master, pady=10)       # Row of buttons
    frame2.pack()
    
    b1 = Button(frame2, text=" Check Inputs ", command=checkInputs)
    b2 = Button(frame2, text=" Start allotment", command=run, state=DISABLED)
    b3 = Button(frame2, text=" Statistics ", command=stats, state=DISABLED)
    b4 = Button(frame2, text=" Clear ", command=clearDisp, state=DISABLED)  
    b1.pack(side=LEFT); b2.pack(side=LEFT)
    b3.pack(side=LEFT); b4.pack(side=LEFT)

    frame3 = Frame(master)       # select of names
    frame3.pack()
    scroll = Scrollbar(frame3, orient=VERTICAL)
    
    lb = Listbox(frame3, yscrollcommand=scroll.set, width=60, height=15)
    scroll.config (command=lb.yview)
    scroll.pack(side=RIGHT, fill=Y)
    lb.pack(side=LEFT,  fill=BOTH, expand=1)
    return master
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)

def run():
  try:
    global btn2
    L1.pack_forget()
    lb.delete(0,END)    
    lb.insert(END, "Initializing the process...\n")
    load_eligible_progs()
    lb.insert(END, "Eligible programmes loaded")
    load_vacancies()
    lb.insert(END, "Old vacancies and additional seats loaded")
    lb.insert(END, "Filtering eligible candidates from the entire applied candidated list...")
    eligible_candidates_list()
    lb.insert(END, "Eligible candidates list prepared and saved to 'EligibleStudents.xlsx'")
    load_student_details()
    lb.insert(END, "Candidates details and data loaded")
    programme_demand_ratings()
    lb.insert(END, "Starting the allotment process...")
    allotment()
    lb.insert(END, "Allotment process complete")
    lb.insert(END, "Printing allotment result to 'BranchChange_result.xlsx'")
    print_result_to_excel()
    lb.insert(END, "Printing complete.")
    lb.insert(END,"Finished.")
    b1.config (state=DISABLED)
    b2.config (state=DISABLED)
    tkMessageBox.showinfo(
            "Info Message",
            "Allotment successfully finished.\nCheckout output at 'OutputFiles' folder"
    )
    btn2 = 1
    b3.config (state=ACTIVE)
    b4.config (state=ACTIVE)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)

def programmeRatings():
  try:
    L7.config(text="Note: These values are extracted from the \n          dept. preference order of the eligible candidates")
    lb2.delete(0, END)
    L6.config( text="Programme Ranking On The Basis Of \nNumber Of Applied Candidates \n", font=("Helvetica", 14), justify=LEFT)
      
    s_list = sorted(programme_ratings.items(), key=operator.itemgetter(1), reverse=True)
    for item in s_list:
      data = "                                                 "+ item[0] +"                    [ "+ str(item[1]) +" ]"
      lb2.insert(END, data)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)

def allotmentStatus():
  try:
    L7.config(text="")
    lb2.delete(0, END)
      
    L6.config( text="Allotment status of all the eligible \ncandidates for branch change programme \n", font=("Helvetica", 14), justify=LEFT )
    s_list = sorted(allotment_status.items(), key=operator.itemgetter(0))
    _c = 1
    lb2.insert(END, "     Enrno.               Allotment status                  Candidate's name")
    lb2.insert(END, "                                                                                                                    ")
    for item in s_list:
      data = "   "+str(item[0]) +"                  "+ item[1][1]  +"                   "+ item[1][0]
      lb2.insert(END, data)
      _c += 1
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)

def IneligibleCandidatesList():
  try:
    L7.pack()
    lb2.delete(0, END)
    L6.config( text="Candidates who are applied but are \nineligible to take part in branch change programme\n", font=("Helvetica", 14), justify=LEFT )
    L7.config( text="Note: 'ID' -  Indicates Penality for indiscipline, \n            'IEC' -  Insufficient earned credits,                  \n" \
                                   "      'CBNS' - Candidate's branch not suitable\n", font=("Helvetica", 10))
    s_list = sorted(ineligible_candidates.items(), key=operator.itemgetter(0))
    _c = 1
    lb2.insert(END, "     Enrno.            Reason for ineligibility                    Candidate's name")
    lb2.insert(END, "                                                                                                                              ")
    for item in s_list:
      data = "   "+str(item[0]) +"       "+ item[1][1]  +"       "+ item[1][0]
      lb2.insert(END, data)
      _c += 1
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)

def remainingVacancies():
  try:
    L7.config(text="")
    lb2.delete(0, END)
    L6.config( text="List of unfilled vacancies in each \ndepartment after the allotment process\n", font=("Helvetica", 14), justify=LEFT )
    #L7.config( text="Note: 'ID/IEC' indicates Both Penality for indiscipline \n        and also having insufficient earned credits", font=("Helvetica", 10))
    s_list = sorted(OV_list.items(), key=operator.itemgetter(0))
    lb2.insert(END, "      Dept.                  GE                   OBC                    SC                  ST")
    lb2.insert(END, "                                                                                                                     ")
    for item in s_list:
      _tmp =""
      if len(item[0]) == 2:     # Just for formatting the data for good apperance
        _tmp= str(item[0]) + "  "
        data = "      "+ _tmp +"                       "+ str(item[1][0])  +"                       "+ str(item[1][1]) +"                       "+ str(item[1][2]) +"                       "+ str(item[1][3])
      else:
        data = "      "+ str(item[0]) +"                       "+ str(item[1][0])  +"                       "+ str(item[1][1]) +"                       "+ str(item[1][2]) +"                       "+ str(item[1][3])
      lb2.insert(END, data)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)
    
  """
  def IneligibleCandidatesList():
  try:
    L7.pack()
    lb2.delete(0, END)
    L6.config( text="Candidates applied but are ineligible \nto take part in branch change programme\n", font=("Helvetica", 14), justify=LEFT )
    L7.config( text="Note: 'ID/IEC' indicates Both Penality for indiscipline \n        and also having insufficient earned credits", font=("Helvetica", 10))
    s_list = sorted(ineligible_candidates.items(), key=operator.itemgetter(0))
    for item in s_list:
      data = str(item[0]) +"     "+ item[1][1]  +"     "+ item[1][0]
      lb2.insert(END, data)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)
  """
def stats():
  try:
    global no_of_applied_students, no_of_eligible_students, lb2, L6, L7, pr_call_no
    win2 = Toplevel()
    frame1 = Frame(win2)       # Contains 3 labels, 2 Buttons
    frame1.pack()
    L6 = Label(frame1, text="Programme Ranking On The Basis Of \nNumber Of Applied Candidates \n", font=("Helvetica", 14))
    L6.pack()
    Label(frame1, text="  Total students applied: "+ str(no_of_applied_students), font=("Helvetica", 10), justify=LEFT).pack()
    Label(frame1, text="  Eligible students: "+ str(no_of_eligible_students), font=("Helvetica", 10),  justify=LEFT).pack()
    _total_alloted_no = len(final_alloted_list_GEN) + len(final_alloted_list_OBC) + len(final_alloted_list_SC) + len(final_alloted_list_ST)
    Label(frame1, text="  Alloted students: "+ str(_total_alloted_no)+ "\n", font=("Helvetica", 10),  justify=LEFT).pack()
    L7 = Label(frame1, text="Note: These values are extracted from the \n          dept. preference order of the eligible candidates", fg="red", font=("Helvetica", 10))
    L7.pack()
		
		# Buttons set
    btn1 = Button(frame1, text=" Dept. ratings ", command=programmeRatings)
    btn2 = Button(frame1, text=" Allotment status ", command=allotmentStatus)
    btn3 = Button(frame1, text=" Ineligible candidates ", command=IneligibleCandidatesList)
    btn4 = Button(frame1, text=" Unfilled vacancies ", command=remainingVacancies)
    btn1.pack(side=LEFT); btn2.pack(side=LEFT);
    btn3.pack(side=LEFT); btn4.pack(side=LEFT); 
    
    frame2 = Frame(win2, pady=10)       # Contains List Box (lb2), Scrollbar
    frame2.pack()
    scroll = Scrollbar(frame2, orient=VERTICAL)
    lb2 = Listbox(frame2, yscrollcommand=scroll.set, width=80, height=15)
    scroll.config (command=lb2.yview)
    scroll.pack(side=RIGHT, fill=Y)
    lb2.pack(side=LEFT,  fill=BOTH, expand=1)
    if pr_call_no == 0:       # This 'if' condition is just to make call to the fn. 'programme_demand_ratings()' only once in the programme lifetime.
      programme_demand_ratings()
      pr_call_no = 1
    s_list = sorted(programme_ratings.items(), key=operator.itemgetter(1), reverse=True)
    for item in s_list:
      data = "                                                 "+ item[0] +"                    [ "+ str(item[1]) +" ]"
      lb2.insert(END, data)
  except Exception as e:
    tkMessageBox.showerror(
            "Error Message",
            "Error occured !! \n(%s)" % str(e)
    )        
    sys.exit(1)
  
  #Vstr = StringVar()	# Variable string which can be updated
  #tk_label = Label(master, textvariable=Vstr, text="Helvetica", font=("Helvetica", 12), justify=LEFT)

def  clearDisp():
  lb.delete(0, END)

  
if __name__ == "__main__":
  win = index()
  win.mainloop()
